from tempfile import template
import openpyxl

def find_row_number(ws,column_number,key):
    row_number = 0
    
    for row in ws.iter_rows():
        row_number +=1
        所属 = row[column_number].value
        if 所属 == None:
            continue
        所属 = 所属.replace('　', '')
        #print(row_number,所属)
        if key == 所属:
            return row_number 
    
def copy_sheet(book,src,dst):
    src_sheet = book[src]
    book.copy_worksheet(src_sheet)
    sheet = book[src+' Copy']
    sheet.title = dst
    return sheet

def generate_monthly_sheet(book,sheet_name):
    if sheet_name in book.sheetnames:
        raise Exception('同じシート名があります')
    sheet = copy_sheet(book,'テンプレ',sheet_name)
    sheet['D17']=sheet['D17'].value.replace('month',sheet_name)
    sheet['J17']=sheet['J17'].value.replace('month',sheet_name)
    return sheet
    
def write_状況表(損傷リスト,file_name,sheet_name):
    # ブックを取得
    book = openpyxl.load_workbook(file_name)
    sheet = generate_monthly_sheet(book,sheet_name)
    print(book.sheetnames)
    # シートを取得
    # セルへ書き込む
    #sheet['D19'] = 損傷リスト['浦和']
    for key in 損傷リスト:
        row_number = find_row_number(sheet,1,key) #1はB列のこと
        if row_number == None:
            continue
        cell_number = 'D'+str(row_number)
        sheet[cell_number] = 損傷リスト[key]['count']
        cell_number = 'G'+str(row_number)
        sheet[cell_number] = 損傷リスト[key]['money']
        #print(row_number)
    for key in 損傷リスト:
        row_number = find_row_number(sheet,23,key) #23はX列のこと
        if row_number == None:
            continue
        cell_number = 'Z'+str(row_number)
        sheet[cell_number] = 損傷リスト[key]['count']
        cell_number = 'AC'+str(row_number)
        sheet[cell_number] = 損傷リスト[key]['money']
        
    # 保存する
    book.save(file_name)
